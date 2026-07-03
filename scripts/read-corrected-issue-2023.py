import argparse
import json
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-file", required=True)
    parser.add_argument("--warning-file", required=True)
    parser.add_argument("--reader", required=True)
    args = parser.parse_args()

    output = subprocess.check_output([
        sys.executable,
        args.reader,
        "--file",
        args.source_file,
    ])
    payload = json.loads(output.decode("utf-8", errors="replace"))
    corrections = read_corrections(Path(args.warning_file))

    rows = []
    for row in payload["rows"]:
        corrected = apply_corrections(row, corrections.get(row["normalized"].get("issue_key"), {}))
        rows.append(corrected)

    print(json.dumps({
        "source_file": args.source_file,
        "warning_file": args.warning_file,
        "physical_row_count": payload.get("physical_row_count"),
        "rows": rows,
    }, ensure_ascii=True))


def read_corrections(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Warning Detail"]
    headers = {str(ws.cell(1, col).value).strip(): col for col in range(1, ws.max_column + 1) if ws.cell(1, col).value}
    result = {}
    for row_number in range(2, ws.max_row + 1):
        issue_key = text(ws.cell(row_number, headers["Issue Key"]).value)
        warning = text(ws.cell(row_number, headers["Warning"]).value)
        correction = text_multiline(ws.cell(row_number, headers["Correction"]).value)
        if not issue_key or not warning or not correction:
            continue
        result.setdefault(issue_key, {})[warning] = correction
    return result


def apply_corrections(row, corrections):
    data = dict(row["normalized"])
    original_issue_key = data.get("issue_key")
    date_start = data.get("create_issue_date")
    issue_key = normalize_legacy_issue_key(original_issue_key, date_start)
    issue_no, sub_issue_no = split_issue_key(issue_key)
    status = data.get("issue_status") or "open"
    is_cancelled = bool(row.get("is_cancelled"))
    cancel_reason = row.get("cancel_reason")
    requester_names = split_people(data.get("requester"))
    abaper_names = []
    created_source = "excel_2023_date_start" if valid_date(date_start) else None
    e070_fallback_requested = False

    for warning, correction in corrections.items():
        normalized = correction.lower()
        if correction == "Tidak dibatalkan":
            is_cancelled = False
            cancel_reason = None
            if status == "cancelled":
                status = "open"
        elif correction == "Tandai Issue dibatalkan":
            is_cancelled = True
            status = "cancelled"
            cancel_reason = cancel_reason or "cancelled"
        elif correction.startswith("Koreksi date start:"):
            date_start = correction.split(":", 1)[1].strip()
            created_source = "excel_2023_correction"
        elif correction == "Ikuti table SAP sekarang (E070)":
            e070_fallback_requested = True
        elif "requester ada" in normalized or "abaper" in normalized:
            parsed_requesters, parsed_abapers, same_as_requester = parse_people_correction(correction)
            if parsed_requesters:
                requester_names = parsed_requesters
            if parsed_abapers:
                abaper_names = parsed_abapers
            elif same_as_requester:
                abaper_names = list(requester_names)

    if is_cancelled:
        status = "cancelled"

    data.update({
        "original_issue_key": original_issue_key,
        "issue_key": issue_key,
        "issue_no": issue_no,
        "sub_issue_no": sub_issue_no,
        "issue_status": status,
        "create_issue_date": date_start if valid_date(date_start) else None,
        "create_issue_date_source": created_source,
        "e070_fallback_requested": e070_fallback_requested,
        "requester_names": requester_names,
        "abaper_names": abaper_names,
        "transport_requests": data.get("transport_requests") or extract_transport_requests(data.get("cr_no")),
    })

    return {
        "source_sheet": row["sheet"],
        "source_row": row["source_row"],
        "occurrence_count": row.get("occurrence_count", 1),
        "is_cancelled": is_cancelled,
        "cancel_reason": cancel_reason if is_cancelled else None,
        "corrections": corrections,
        "raw": row.get("raw", {}),
        "data": data,
    }


def normalize_legacy_issue_key(issue_key, date_start):
    if not issue_key:
        return issue_key
    match = re.match(r"^(80|90)(\d+)-(\d+)$", str(issue_key))
    if not match:
        return issue_key
    year = str(date_start or "")[:4]
    if match.group(1) == "80" and year == "2018":
        return f"18{match.group(2)}-{match.group(3)}"
    if match.group(1) == "90" and year == "2019":
        return f"19{match.group(2)}-{match.group(3)}"
    return issue_key


def split_issue_key(issue_key):
    match = re.match(r"^(\d+)-(\d+)$", str(issue_key or ""))
    if not match:
        return None, None
    return int(match.group(1)), match.group(2).zfill(2)


def parse_people_correction(correction):
    requesters = []
    abapers = []
    same_as_requester = "ABAPer disamakan dengan Requester" in correction
    for line in [item.strip() for item in correction.splitlines() if item.strip()]:
        if line.lower().startswith("requester ada") and ":" in line:
            requesters = split_people(line.split(":", 1)[1])
        elif line.lower().startswith("abaper:"):
            abapers = split_people(line.split(":", 1)[1])
    return requesters, abapers, same_as_requester


def split_people(value):
    clean = text(value)
    if not clean:
        return []
    return [item.strip() for item in re.split(r"\s*/\s*|\s*,\s*|\s+dan\s+", clean, flags=re.I) if item.strip()]


def extract_transport_requests(value):
    clean = text(value)
    if not clean:
        return []
    return sorted(set(item.upper() for item in re.findall(r"\bTR[A-Z0-9]{6,}\b", clean, flags=re.I)))


def valid_date(value):
    return bool(re.match(r"^\d{4}-\d{2}-\d{2}$", str(value or "")))


def text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def text_multiline(value):
    if value is None:
        return ""
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in str(value).strip().splitlines()]
    return "\n".join(line for line in lines if line)


if __name__ == "__main__":
    main()
