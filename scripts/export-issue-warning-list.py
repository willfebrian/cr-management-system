import os
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def main():
    out_path = os.environ["OUT_XLSX"]
    input_json = os.environ["WARNINGS_JSON"]
    with open(input_json, "r", encoding="utf-8") as file:
        payload = json.load(file)
    batch = payload["batch"]
    rows = payload["rows"]

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_list = wb.create_sheet("Warning List")
    ws_detail = wb.create_sheet("Warning Detail")

    styles = make_styles()
    write_summary(ws_summary, batch, styles)
    warning_counts = write_warning_list(ws_list, rows, styles)
    write_warning_detail(ws_detail, rows, warning_counts, styles)
    write_warning_summary(ws_summary, warning_counts, styles)

    wb.save(out_path)
    print(out_path)


def read_warning_rows(conn, batch_id):
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, source_file, sheet_name, import_mode, status, total_rows, valid_rows,
                   imported_rows, warning_count, error_count, summary, started_at, finished_at
            FROM cr_management.issue_import_batches
            WHERE id = %s
            """,
            (batch_id,),
        )
        batch = cur.fetchone()
        if not batch:
            raise SystemExit(f"Batch {batch_id} not found")

        cur.execute(
            """
            SELECT row_number, issue_key, issue_no, sub_issue_no, row_status, is_cancelled,
                   cancel_reason, raw_data, normalized_data, warnings, errors
            FROM cr_management.issue_import_rows
            WHERE batch_id = %s
              AND cardinality(warnings) > 0
            ORDER BY row_number
            """,
            (batch_id,),
        )
        rows = cur.fetchall()
    return batch, rows


def make_styles():
    thin = Side(style="thin", color="D9E2F3")
    return {
        "header_fill": PatternFill("solid", fgColor="1F4E78"),
        "header_font": Font(color="FFFFFF", bold=True),
        "warn_fill": PatternFill("solid", fgColor="FFF2CC"),
        "cancel_fill": PatternFill("solid", fgColor="FCE4D6"),
        "label_fill": PatternFill("solid", fgColor="D9EAF7"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def write_summary(ws, batch, styles):
    summary_labels = [
        ("Batch ID", batch["id"]),
        ("Source File", batch["source_file"]),
        ("Sheet Name", batch["sheet_name"]),
        ("Import Mode", batch["import_mode"]),
        ("Status", batch["status"]),
        ("Total Rows", batch["total_rows"]),
        ("Valid Rows", batch["valid_rows"]),
        ("Imported Rows", batch["imported_rows"]),
        ("Warning Count", batch["warning_count"]),
        ("Error Count", batch["error_count"]),
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for row_index, (label, value) in enumerate(summary_labels, start=1):
        ws.cell(row_index, 1, label)
        ws.cell(row_index, 2, value)
        ws.cell(row_index, 1).font = Font(bold=True)
        ws.cell(row_index, 1).fill = styles["label_fill"]
        ws.cell(row_index, 1).border = styles["border"]
        ws.cell(row_index, 2).border = styles["border"]
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90


def write_warning_list(ws, rows, styles):
    headers = [
        "Row No",
        "Source Sheet",
        "Source Row",
        "Issue Key",
        "Issue Name",
        "Status",
        "CR No",
        "Requester",
        "Date Start",
        "Is Cancelled",
        "Cancel Reason",
        "Warnings",
        "Errors",
    ]
    ws.append(headers)
    style_header(ws[1], styles)

    warning_counts = {}
    for dbrow in rows:
        row_number = dbrow["row_number"]
        issue_key = dbrow["issue_key"]
        is_cancelled = dbrow["is_cancelled"]
        cancel_reason = dbrow["cancel_reason"]
        norm = dbrow.get("normalized_data") or {}
        raw = dbrow.get("raw_data") or {}
        warnings = dbrow.get("warnings") or []
        errors = dbrow.get("errors") or []
        for warning in warnings:
            warning_counts[warning] = warning_counts.get(warning, 0) + 1

        ws.append(
            [
                row_number,
                raw.get("source_sheet"),
                raw.get("source_row"),
                issue_key,
                norm.get("issue_name"),
                norm.get("issue_status"),
                norm.get("cr_no"),
                norm.get("requester"),
                norm.get("create_issue_date"),
                "Yes" if is_cancelled else "No",
                cancel_reason,
                "\n".join(warnings),
                "\n".join(errors),
            ]
        )
        fill = styles["cancel_fill"] if is_cancelled else styles["warn_fill"]
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = styles["border"]

    widths = [10, 14, 10, 14, 55, 12, 18, 20, 14, 14, 24, 55, 35]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    return warning_counts


def write_warning_detail(ws, rows, warning_counts, styles):
    ws.append(["Warning", "Count", "Issue Key", "Source", "Issue Name", "CR No"])
    style_header(ws[1], styles)
    for dbrow in rows:
        issue_key = dbrow["issue_key"]
        norm = dbrow.get("normalized_data") or {}
        raw = dbrow.get("raw_data") or {}
        warnings = dbrow.get("warnings") or []
        for warning in warnings or []:
            ws.append(
                [
                    warning,
                    warning_counts.get(warning, 0),
                    issue_key,
                    f"{raw.get('source_sheet')}:{raw.get('source_row')}",
                    norm.get("issue_name"),
                    norm.get("cr_no"),
                ]
            )
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = styles["border"]

    for col, width in enumerate([55, 10, 14, 14, 55, 18], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def write_warning_summary(ws, warning_counts, styles):
    start = 14
    ws.cell(start, 1, "Warning Type")
    ws.cell(start, 2, "Count")
    style_header(ws[start], styles)
    for row_index, (warning, count) in enumerate(sorted(warning_counts.items(), key=lambda item: (-item[1], item[0])), start=start + 1):
        ws.cell(row_index, 1, warning)
        ws.cell(row_index, 2, count)
        ws.cell(row_index, 1).border = styles["border"]
        ws.cell(row_index, 2).border = styles["border"]


def style_header(row, styles):
    for cell in row:
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = styles["border"]


if __name__ == "__main__":
    main()
