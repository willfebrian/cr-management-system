import argparse
import json
import re
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


SHEET_ORDER = {
    "2022": 0,
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "May": 5,
    "June": 6,
    "Jul": 7,
    "Aug": 8,
    "Sep": 9,
    "Okt": 10,
    "Nov": 11,
    "Dec": 12,
}

FIELD_MAP = {
    "seq. no.": "sequence_id",
    "issue no.": "issue_key",
    "issue no": "issue_key",
    "date start": "create_issue_date",
    "issue": "issue_name",
    "status": "source_status",
    "progress (%)": "progress",
    "date closed": "closed_date",
    "application / program name": "program_function_object",
    "cr. no.": "cr_no",
    "cr no.": "cr_no",
    "modul": "module",
    "requester": "requester",
    "group category": "group_category",
    "task category": "task_category",
    "cr description": "cr_description",
    "activity description": "activity_description",
}


def normalize_header(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def normalize_value(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        return value.strip()
    return value


def text_or_none(value):
    if value is None or value == "":
        return None
    text = re.sub(r"\s+", " ", str(value).strip())
    return text or None


def split_issue_key(value):
    text = text_or_none(value)
    if not text:
        return None, None, None
    match = re.match(r"^(\d+)-(\d+)$", text)
    if not match:
        return None, None, text
    return int(match.group(1)), match.group(2).zfill(2), text


def normalize_status(value):
    text = text_or_none(value)
    if not text:
        return "open"
    lowered = text.lower()
    if lowered in {"finish", "finished", "ok"}:
        return "ok"
    if lowered in {"cancelled", "canceled"}:
        return "cancelled"
    if lowered in {"in progress", "testing"}:
        return "open"
    return lowered.replace(" ", "_")


def extract_transport_requests(value):
    text = text_or_none(value)
    if not text:
        return []
    return sorted(set(item.upper() for item in re.findall(r"\bTR[A-Z0-9]{6,}\b", text, flags=re.I)))


def is_red_cell(cell):
    if not cell.fill or not cell.fill.fgColor:
        return False
    color = cell.fill.fgColor
    if color.type != "rgb" or not color.rgb:
        return False
    return color.rgb.upper() in {"FFFF0000", "00FF0000", "FFFFC7CE", "FFFF9999", "FFFF6666", "FFFF8080"}


def row_score(row):
    data = row["normalized"]
    filled = sum(1 for value in data.values() if value not in (None, ""))
    sheet_score = SHEET_ORDER.get(row["sheet"], 99)
    return (sheet_score, filled, row["source_row"])


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True)
    parser.add_argument("--header-row", type=int, default=4)
    parser.add_argument("--data-row", type=int, default=5)
    args = parser.parse_args()

    workbook_path = Path(args.file)
    wb = load_workbook(workbook_path, data_only=True)
    physical_rows = []
    by_issue = {}
    occurrences = {}

    for ws in wb.worksheets:
        headers = {}
        unnamed_formula_columns = []
        for col in range(1, ws.max_column + 1):
            label = normalize_header(ws.cell(args.header_row, col).value)
            if not label:
                unnamed_formula_columns.append(col)
                continue
            field = FIELD_MAP.get(label)
            if field:
                headers[col] = field

        for source_row in range(args.data_row, ws.max_row + 1):
            raw = {}
            normalized = {}
            for col, field in headers.items():
                value = normalize_value(ws.cell(source_row, col).value)
                raw[field] = value
                if value not in (None, ""):
                    normalized[field] = value

            if not any(normalized.get(key) for key in ("issue_key", "issue_name", "cr_no")):
                continue

            issue_no, sub_issue_no, issue_key = split_issue_key(normalized.get("issue_key"))
            normalized["issue_no"] = issue_no
            normalized["sub_issue_no"] = sub_issue_no
            normalized["issue_key"] = issue_key
            normalized["issue_status"] = normalize_status(normalized.get("source_status"))
            normalized["transport_requests"] = extract_transport_requests(normalized.get("cr_no"))

            is_cancelled = normalized["issue_status"] == "cancelled" or any(
                is_red_cell(ws.cell(source_row, col)) for col in range(1, min(ws.max_column, 46) + 1)
            )
            cancel_reason_values = []
            for col in range(47, ws.max_column + 1):
                value = normalize_value(ws.cell(source_row, col).value)
                if value not in (None, ""):
                    cancel_reason_values.append(str(value))

            row = {
                "sheet": ws.title,
                "source_row": source_row,
                "raw": raw,
                "normalized": normalized,
                "is_cancelled": is_cancelled,
                "cancel_reason": " | ".join(cancel_reason_values) if cancel_reason_values else None,
                "excluded_formula_columns": unnamed_formula_columns,
            }
            physical_rows.append(row)
            key = issue_key or f"{ws.title}:{source_row}"
            occurrences[key] = occurrences.get(key, 0) + 1
            if key not in by_issue or row_score(row) >= row_score(by_issue[key]):
                by_issue[key] = row

    rows = []
    for index, row in enumerate(by_issue.values(), start=1):
        issue_key = row["normalized"].get("issue_key")
        row["staging_row_number"] = index
        row["occurrence_count"] = occurrences.get(issue_key, 1)
        rows.append(row)

    rows.sort(key=lambda item: (
        item["normalized"].get("issue_no") or 99999999,
        item["normalized"].get("sub_issue_no") or "99",
        item["sheet"],
        item["source_row"],
    ))

    print(json.dumps({
        "file": str(workbook_path),
        "sheets": [ws.title for ws in wb.worksheets],
        "physical_row_count": len(physical_rows),
        "rows": rows,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
