import argparse
import json
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


FIELD_MAP = {
    "Sequence ID": "sequence_id",
    "Issue No": "issue_no",
    "Sub Issue No.": "sub_issue_no",
    "Issue Name": "issue_name",
    "CR Helpdesk Form": "cr_helpdesk_form",
    "CR Helpdesk No.": "cr_helpdesk_no",
    "Requester": "requester",
    "Problem Analysis": "problem_analysis",
    "Impact Analysis": "impact_analysis",
    "ABAPer": "abaper",
    "Email Subject": "email_subject",
    "Email Date Received": "email_date_received",
    "Create Issue Date": "create_issue_date",
    "GLPI Ticket Number": "glpi_ticket_number",
    "CR No.": "cr_no",
    "CR Description": "cr_description",
    "Program/Function/Object": "program_function_object",
    "Date Tested (DEV)": "dev_tested_date",
    "Tester (DEV)": "dev_tester",
    "Date Evaluated (DEV)": "dev_evaluated_date",
    "Evaluator (DEV)": "dev_evaluator",
    "Status": "status",
    "Transport By (QA)": "transported_by_qa",
    "Date Tested (QA)": "qa_tested_date",
    "Tester (QA)": "qa_tester",
    "Date Evaluated (QA)": "qa_evaluated_date",
    "Evaluator (QA)": "qa_evaluator",
    "Requester (PRD)": "prd_requester",
    "Request Date (PRD)": "prd_requested_date",
    "Evaluator (PRD)": "prd_evaluator",
    "Evaluated Date (PRD)": "prd_evaluated_date",
    "Approval": "approval",
    "Approval Date": "approval_date",
    "Executor": "executor",
}


def normalize_value(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        return value.strip()
    return value


def is_red_cell(cell):
    if not cell.fill or not cell.fill.fgColor:
        return False
    color = cell.fill.fgColor
    if color.type != "rgb" or not color.rgb:
        return False
    return color.rgb.upper() in {"FFFF0000", "00FF0000"}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True)
    parser.add_argument("--sheet", default="INPUT")
    parser.add_argument("--header-row", type=int, default=4)
    parser.add_argument("--data-row", type=int, default=5)
    args = parser.parse_args()

    workbook_path = Path(args.file)
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[args.sheet]

    headers = {}
    duplicate_counts = {}
    for col in range(1, ws.max_column + 1):
        label = ws.cell(args.header_row, col).value
        if not label:
            continue
        label = str(label).strip()
        field = FIELD_MAP.get(label)
        if not field:
            continue
        if field == "status":
            duplicate_counts[field] = duplicate_counts.get(field, 0) + 1
            field = "dev_status" if duplicate_counts["status"] == 1 else "qa_status"
        headers[col] = field

    rows = []
    for row_number in range(args.data_row, ws.max_row + 1):
        raw = {}
        normalized = {}
        for col, field in headers.items():
            value = normalize_value(ws.cell(row_number, col).value)
            raw[field] = value
            if value not in (None, ""):
                normalized[field] = value

        issue_no = normalized.get("issue_no")
        issue_name = normalized.get("issue_name")
        if issue_no in (None, "") and issue_name in (None, ""):
            continue

        is_cancelled = any(is_red_cell(ws.cell(row_number, col)) for col in range(1, min(ws.max_column, 46) + 1))
        cancel_reason_values = []
        for col in range(47, ws.max_column + 1):
            value = normalize_value(ws.cell(row_number, col).value)
            if value not in (None, ""):
                cancel_reason_values.append(str(value))
        cancel_reason = " | ".join(cancel_reason_values) if cancel_reason_values else None

        rows.append({
            "row_number": row_number,
            "raw": raw,
            "normalized": normalized,
            "is_cancelled": is_cancelled,
            "cancel_reason": cancel_reason,
        })

    print(json.dumps({
        "file": str(workbook_path),
        "sheet": args.sheet,
        "rows": rows,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
